from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.core.exceptions import ValidationError
from django.utils import timezone
from decimal import Decimal, InvalidOperation
import pandas as pd
import openpyxl
import csv
import io
from datetime import datetime

from ..models import Qualified_for_ste
from enrollment_app.models import Student


def get_user_avatar_url(user):
    """Generate avatar URL or initials for user"""
    try:
        if hasattr(user, 'profile') and user.profile.photo:
            return user.profile.photo.url
    except:
        pass
    return None


def get_user_initials(user):
    """Get user initials from first and last name"""
    first_initial = user.first_name[0].upper() if user.first_name else ''
    last_initial = user.last_name[0].upper() if user.last_name else ''
    return f"{first_initial}{last_initial}"


def is_ste_coordinator(user):
    """Check if user is an STE coordinator"""
    if not hasattr(user, 'profile'):
        return False
    profile = user.profile
    return profile and profile.program and profile.program.code == 'STE'


def ste_access_denied_response(request, is_ajax=False):
    """Return appropriate response for non-STE access attempt"""
    if is_ajax:
        return JsonResponse({
            'success': False,
            'message': 'Upload Results is only available for STE coordinators.'
        }, status=403)
    messages.error(request, 'Upload Results is only available for STE coordinators.')
    return redirect('coordinator:dashboard')


@login_required
def results_upload(request):
    """Main results upload page"""
    # Check if user is STE coordinator - this feature is only for STE program
    if not is_ste_coordinator(request.user):
        return ste_access_denied_response(request)
    
    # Get user information
    user = request.user
    user_profile_obj = user.profile if hasattr(user, 'profile') else None
    
    user_profile = {
        'full_name': f"{user.last_name}, {user.first_name}",
        'first_name': user.first_name,
        'last_name': user.last_name,
        'initials': get_user_initials(user),
        'role': user_profile_obj.get_user_type_display() if user_profile_obj else 'Coordinator',
        'user_type': user_profile_obj.user_type if user_profile_obj else 'coordinator',
        'program': user_profile_obj.get_program_name() if user_profile_obj else 'N/A',
        'avatar': get_user_avatar_url(user),
        'position': user_profile_obj.get_position_name() if user_profile_obj else 'N/A',
        'department': user_profile_obj.get_department_name() if user_profile_obj else 'N/A',
    }
    
    # Get recent uploads
    recent_uploads = Qualified_for_ste.objects.select_related('updated_by').order_by('-updated_at')[:10]
    
    # Get statistics
    stats = {
        'total_records': Qualified_for_ste.objects.count(),
        'qualified': Qualified_for_ste.objects.filter(status='qualified').count(),
        'pending': Qualified_for_ste.objects.filter(status='pending').count(),
        'not_qualified': Qualified_for_ste.objects.filter(status='not_qualified').count(),
    }
    
    context = {
        'user_profile': user_profile,
        'user_initials': get_user_initials(user),
        'recent_uploads': recent_uploads,
        'stats': stats,
        'program_code': user_profile_obj.program.code if user_profile_obj and user_profile_obj.program else None,
    }
    
    return render(request, 'coordinator_app/resultsUpload.html', context)


@login_required
@require_http_methods(["POST"])
def manual_entry(request):
    """Handle manual entry of student results"""
    # Check if user is STE coordinator
    if not is_ste_coordinator(request.user):
        return ste_access_denied_response(request, is_ajax=True)
    
    try:
        # Get form data
        student_lrn = request.POST.get('student_lrn', '').strip()
        exam_score = request.POST.get('exam_score', '').strip()
        interview_score = request.POST.get('interview_score', '').strip()
        status = request.POST.get('status', 'pending').strip()
        remarks = request.POST.get('remarks', '').strip()
        
        # Validate required fields
        if not all([student_lrn, exam_score, interview_score]):
            return JsonResponse({
                'success': False,
                'message': 'All fields are required'
            }, status=400)
        
        # Validate LRN format (should be 12 digits)
        if not student_lrn.isdigit() or len(student_lrn) != 12:
            return JsonResponse({
                'success': False,
                'message': 'Invalid LRN format. LRN must be 12 digits'
            }, status=400)
        
        # Note: Student doesn't need to exist in Student model yet
        # Results can be uploaded before student enrollment
        
        # Validate scores
        try:
            exam_score = Decimal(exam_score)
            interview_score = Decimal(interview_score)
            
            if not (0 <= exam_score <= 100):
                raise ValueError("Exam score must be between 0 and 100")
            if not (0 <= interview_score <= 100):
                raise ValueError("Interview score must be between 0 and 100")
        except (ValueError, InvalidOperation) as e:
            return JsonResponse({
                'success': False,
                'message': f'Invalid score: {str(e)}'
            }, status=400)
        
        # Create or update qualification record
        qualification, created = Qualified_for_ste.objects.update_or_create(
            student_lrn=student_lrn,
            defaults={
                'exam_score': exam_score,
                'interview_score': interview_score,
                'status': status,
                'remarks': remarks,
                'updated_by': request.user,
            }
        )
        
        action = 'created' if created else 'updated'
        
        return JsonResponse({
            'success': True,
            'message': f'Student record {action} successfully',
            'data': {
                'lrn': qualification.student_lrn,
                'exam_score': float(qualification.exam_score),
                'interview_score': float(qualification.interview_score),
                'status': qualification.status,
                'total_score': float(qualification.get_total_score()),
                'average_score': float(qualification.get_average_score()),
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def bulk_upload(request):
    """Handle bulk upload of results via Excel/CSV"""
    # Check if user is STE coordinator
    if not is_ste_coordinator(request.user):
        return ste_access_denied_response(request, is_ajax=True)
    
    try:
        # Check if file was uploaded
        if 'file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': 'No file uploaded'
            }, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Validate file size (10MB limit)
        if uploaded_file.size > 10 * 1024 * 1024:
            return JsonResponse({
                'success': False,
                'message': 'File size exceeds 10MB limit'
            }, status=400)
        
        # Validate file extension
        file_name = uploaded_file.name.lower()
        if not (file_name.endswith('.xlsx') or file_name.endswith('.xls') or file_name.endswith('.csv')):
            return JsonResponse({
                'success': False,
                'message': 'Invalid file format. Please upload .xlsx, .xls, or .csv file'
            }, status=400)
        
        # Parse file based on type
        try:
            if file_name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error reading file: {str(e)}'
            }, status=400)
        
        # Validate required columns
        required_columns = ['student_lrn', 'exam_score', 'interview_score', 'status']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return JsonResponse({
                'success': False,
                'message': f'Missing required columns: {", ".join(missing_columns)}'
            }, status=400)
        
        # Process records
        results = {
            'total': len(df),
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # Validate and clean data
                    student_lrn = str(row['student_lrn']).strip()
                    
                    # Validate LRN
                    if not student_lrn.isdigit() or len(student_lrn) != 12:
                        raise ValidationError(f"Row {index + 2}: Invalid LRN format")
                    
                    # Note: Student doesn't need to exist in Student model yet
                    # Results can be uploaded before student enrollment
                    
                    # Validate scores
                    exam_score = Decimal(str(row['exam_score']))
                    interview_score = Decimal(str(row['interview_score']))
                    
                    if not (0 <= exam_score <= 100):
                        raise ValidationError(f"Row {index + 2}: Exam score must be between 0 and 100")
                    if not (0 <= interview_score <= 100):
                        raise ValidationError(f"Row {index + 2}: Interview score must be between 0 and 100")
                    
                    # Validate status
                    status = str(row['status']).strip().lower()
                    valid_statuses = ['pending', 'qualified', 'not_qualified', 'waitlisted']
                    if status not in valid_statuses:
                        status = 'pending'
                    
                    # Get optional remarks
                    remarks = str(row.get('remarks', '')).strip() if pd.notna(row.get('remarks')) else ''
                    
                    # Create or update record
                    Qualified_for_ste.objects.update_or_create(
                        student_lrn=student_lrn,
                        defaults={
                            'exam_score': exam_score,
                            'interview_score': interview_score,
                            'status': status,
                            'remarks': remarks,
                            'updated_by': request.user,
                        }
                    )
                    
                    results['success'] += 1
                    
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"Row {index + 2}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': f'Processing complete. {results["success"]} records imported, {results["failed"]} failed.',
            'data': results
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error processing file: {str(e)}'
        }, status=500)


@login_required
def download_template(request):
    """Download Excel template for bulk upload"""
    # Check if user is STE coordinator
    if not is_ste_coordinator(request.user):
        return ste_access_denied_response(request)
    
    try:
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results Template"
        
        # Add headers
        headers = ['student_lrn', 'exam_score', 'interview_score', 'status', 'remarks']
        ws.append(headers)
        
        # Add sample data
        sample_data = [
            ['123456789012', '85.50', '90.00', 'qualified', 'Excellent performance'],
            ['123456789013', '75.00', '80.50', 'pending', 'Under review'],
            ['123456789014', '65.00', '70.00', 'not_qualified', 'Did not meet requirements'],
        ]
        
        for row in sample_data:
            ws.append(row)
        
        # Style headers
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="results_template_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error generating template: {str(e)}')
        return redirect('coordinator:results_upload')


@login_required
def export_results(request):
    """Export all results to Excel"""
    # Check if user is STE coordinator
    if not is_ste_coordinator(request.user):
        return ste_access_denied_response(request)
    
    try:
        # Get all qualifications
        qualifications = Qualified_for_ste.objects.all().order_by('-updated_at')
        
        # Create DataFrame
        data = []
        for q in qualifications:
            data.append({
                'LRN': q.student_lrn,
                'Exam Score': float(q.exam_score),
                'Interview Score': float(q.interview_score),
                'Total Score': float(q.get_total_score()),
                'Average Score': float(q.get_average_score()),
                'Status': q.get_status_display(),
                'Remarks': q.remarks or '',
                'Updated By': q.updated_by.get_full_name() if q.updated_by else '',
                'Updated At': q.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Results', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Results']
            
            # Style headers
            for cell in worksheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="all_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting results: {str(e)}')
        return redirect('coordinator:results_upload')


@login_required
@require_http_methods(["DELETE"])
def delete_result(request, lrn):
    """Delete a specific result"""
    # Check if user is STE coordinator
    if not is_ste_coordinator(request.user):
        return ste_access_denied_response(request, is_ajax=True)
    
    try:
        qualification = Qualified_for_ste.objects.get(student_lrn=lrn)
        qualification.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Record deleted successfully'
        })
        
    except Qualified_for_ste.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Record not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting record: {str(e)}'
        }, status=500)


@login_required
def view_result(request, lrn):
    """View details of a specific result"""
    # Check if user is STE coordinator
    if not is_ste_coordinator(request.user):
        return ste_access_denied_response(request, is_ajax=True)
    
    try:
        qualification = Qualified_for_ste.objects.get(student_lrn=lrn)
        
        # Try to get student name if exists, otherwise use LRN
        try:
            student = Student.objects.get(lrn=lrn)
            student_name = f"{student.last_name}, {student.first_name} {student.middle_name}"
        except Student.DoesNotExist:
            student_name = f"Student (LRN: {lrn})"
        
        data = {
            'lrn': qualification.student_lrn,
            'student_name': student_name,
            'exam_score': float(qualification.exam_score),
            'interview_score': float(qualification.interview_score),
            'total_score': float(qualification.get_total_score()),
            'average_score': float(qualification.get_average_score()),
            'status': qualification.status,
            'status_display': qualification.get_status_display(),
            'remarks': qualification.remarks or '',
            'updated_by': qualification.updated_by.get_full_name() if qualification.updated_by else '',
            'updated_at': qualification.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
        }
        
        return JsonResponse({
            'success': True,
            'data': data
        })
        
    except Qualified_for_ste.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Record not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving record: {str(e)}'
        }, status=500)