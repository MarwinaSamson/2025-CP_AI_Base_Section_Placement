from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from admin_app.decorators import admin_required
from admin_app.models import SchoolYear, Section, Program, GradeLevel
from enrollment_app.models import (
    Student, StudentEnrollment, StudentAcademicYearStatus, ProgramSelection
)
from coordinator_app.models import AcademicPerformance, ProbationRecord
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import Count, Avg, Q


@admin_required
def archive(request):
    """Main archive page — admin only."""
    active_school_year = SchoolYear.objects.filter(is_active=True).first()
    context = {
        'active_page': 'archive',
        'active_school_year': active_school_year,
    }
    return render(request, 'admin_app/archive.html', context)


@admin_required
def archive_school_years(request):
    """Returns all INACTIVE school years for the archive dropdown."""
    school_years = SchoolYear.objects.filter(
        is_active=False
    ).order_by('-year_label').values('id', 'year_label', 'start_date', 'end_date')

    return JsonResponse({
        'success': True,
        'school_years': [
            {
                'id': sy['id'],
                'year_label': sy['year_label'],
                'start_date': sy['start_date'].strftime('%b %d, %Y') if sy['start_date'] else '',
                'end_date': sy['end_date'].strftime('%b %d, %Y') if sy['end_date'] else '',
            }
            for sy in school_years
        ]
    })


@admin_required
def archive_enrollment(request):
    """Tab 1 — Enrollment summary for a given school year."""
    sy_id = request.GET.get('sy')
    if not sy_id:
        return JsonResponse({'success': False, 'error': 'sy parameter required.'}, status=400)

    try:
        school_year = SchoolYear.objects.get(id=sy_id, is_active=False)
    except SchoolYear.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'School year not found or is currently active.'}, status=404)

    enrollments = StudentEnrollment.objects.filter(school_year=school_year)

    # Overall counts
    total = enrollments.count()
    by_type = {
        'new': enrollments.filter(enrollee_type='new').count(),
        'continuing': enrollments.filter(enrollee_type='continuing').count(),
        'transferee': enrollments.filter(enrollee_type='transferee').count(),
        'returnee': enrollments.filter(enrollee_type='returnee').count(),
    }
    by_status = {
        'approved': enrollments.filter(enrollment_status='approved').count(),
        'rejected': enrollments.filter(enrollment_status='rejected').count(),
        'pending': enrollments.filter(enrollment_status='pending').count(),
        'submitted': enrollments.filter(enrollment_status='submitted').count(),
    }

    # Per program breakdown
    program_selections = ProgramSelection.objects.filter(
        school_year=school_year
    ).values('selected_program_code').annotate(
        total=Count('student'),
        approved=Count('student', filter=Q(admin_approved=True)),
        rejected=Count('student', filter=Q(admin_rejected=True)),
    ).order_by('selected_program_code')

    programs = []
    for p in program_selections:
        if p['selected_program_code']:
            programs.append({
                'code': p['selected_program_code'],
                'total': p['total'],
                'approved': p['approved'],
                'rejected': p['rejected'],
                'pending': p['total'] - p['approved'] - p['rejected'],
            })

    return JsonResponse({
        'success': True,
        'school_year': school_year.year_label,
        'total': total,
        'by_type': by_type,
        'by_status': by_status,
        'by_program': programs,
    })


@admin_required
def archive_sections(request):
    """Tab 2 — Sections and their student masterlists."""
    sy_id = request.GET.get('sy')
    section_id = request.GET.get('section_id')

    if not sy_id:
        return JsonResponse({'success': False, 'error': 'sy parameter required.'}, status=400)

    try:
        school_year = SchoolYear.objects.get(id=sy_id, is_active=False)
    except SchoolYear.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'School year not found.'}, status=404)

    # If section_id provided — return full grade table for that section
    if section_id:
        try:
            section = Section.objects.select_related(
                'program', 'grade_level', 'adviser'
            ).get(id=section_id, school_year=school_year)
        except Section.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Section not found.'}, status=404)

        program_selections = ProgramSelection.objects.filter(
            assigned_section=section,
            admin_approved=True,
        ).select_related('student', 'student__student_data').order_by(
            'student__student_data__last_name'
        )

        # Get subjects for this section
        from admin_app.models import Subject
        from coordinator_app.views.coor_masterlist_views import _resolve_subject_program
        subject_program = _resolve_subject_program(section)
        subjects = list(Subject.objects.filter(
            program=subject_program, is_active=True
        ).order_by('name'))

        students = []
        for ps in program_selections:
            student = ps.student
            sd = getattr(student, 'student_data', None)
            if not sd:
                continue

            # Get grades for this student
            performances = AcademicPerformance.objects.filter(
                student=student,
                school_year=school_year,
                grade_level=section.grade_level,
            ).select_related('subject')

            grade_map = {}
            for perf in performances:
                subj_id = perf.subject_id
                if subj_id not in grade_map:
                    grade_map[subj_id] = {
                        'q1': None, 'q2': None,
                        'q3': None, 'q4': None, 'final': None
                    }
                q_map = {1:'q1', 2:'q2', 3:'q3', 4:'q4', 5:'final'}
                field = q_map.get(perf.quarter)
                if field:
                    grade_map[subj_id][field] = float(perf.grade)

            # Build subject grades list
            subject_grades = []
            final_ratings = []
            for subj in subjects:
                g = grade_map.get(subj.id, {})
                q1 = g.get('q1')
                q2 = g.get('q2')
                q3 = g.get('q3')
                q4 = g.get('q4')
                final = g.get('final')
                if final is None and all(x is not None for x in [q1, q2, q3, q4]):
                    final = round((q1 + q2 + q3 + q4) / 4, 2)
                if final is not None:
                    final_ratings.append(final)
                subject_grades.append({
                    'subject': subj.name,
                    'q1': q1, 'q2': q2, 'q3': q3, 'q4': q4,
                    'final': final,
                    'status': ('Passed' if final >= 75 else 'Failed') if final is not None else None,
                })

            overall = round(sum(final_ratings) / len(final_ratings), 2) if final_ratings else None

            # Get promotion status
            try:
                acad_status = StudentAcademicYearStatus.objects.get(
                    student=student, school_year=school_year
                )
                promotion_status = acad_status.final_status
            except StudentAcademicYearStatus.DoesNotExist:
                promotion_status = None

            students.append({
                'lrn': student.lrn,
                'name': f"{sd.last_name}, {sd.first_name}{' ' + sd.middle_name if sd.middle_name else ''}",
                'gender': sd.gender,
                'subject_grades': subject_grades,
                'overall_grade': overall,
                'promotion_status': promotion_status,
            })

        return JsonResponse({
            'success': True,
            'section': {
                'id': section.id,
                'name': section.name,
                'program': section.program.code if section.program else '',
                'grade_level': section.grade_level.name if section.grade_level else '',
                'adviser': section.adviser.get_full_name() if section.adviser else '—',
            },
            'subjects': [s.name for s in subjects],
            'students': students,
        })

    # Otherwise — return list of all sections for this school year
    sections = Section.objects.filter(
        school_year=school_year
    ).select_related('program', 'grade_level', 'adviser').order_by(
        'program__code', 'grade_level__code', 'name'
    )

    section_list = []
    for s in sections:
        student_count = ProgramSelection.objects.filter(
            assigned_section=s, admin_approved=True
        ).count()
        section_list.append({
            'id': s.id,
            'name': s.name,
            'program': s.program.code if s.program else '',
            'grade_level': s.grade_level.name if s.grade_level else '',
            'adviser': s.adviser.get_full_name() if s.adviser else '—',
            'student_count': student_count,
            'regular_track': s.regular_track or '',
        })

    return JsonResponse({
        'success': True,
        'school_year': school_year.year_label,
        'sections': section_list,
    })


@admin_required
def archive_academic(request):
    """Tab 3 — Academic performance summary."""
    sy_id = request.GET.get('sy')
    if not sy_id:
        return JsonResponse({'success': False, 'error': 'sy parameter required.'}, status=400)

    try:
        school_year = SchoolYear.objects.get(id=sy_id, is_active=False)
    except SchoolYear.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'School year not found.'}, status=404)

    statuses = StudentAcademicYearStatus.objects.filter(
        school_year=school_year
    ).select_related('student', 'student__student_data', 'section', 'section__program', 'grade_level')

    # Aggregate counts
    total = statuses.count()
    promoted = statuses.filter(final_status='promoted').count()
    retained = statuses.filter(final_status='retained').count()
    incomplete = statuses.filter(final_status='pending').count()
    others = total - promoted - retained - incomplete

    # Per program breakdown
    program_stats = {}
    section_stats = {}

    for s in statuses:
        # Program grouping
        program_code = s.section.program.code if s.section and s.section.program else 'Unknown'
        if program_code not in program_stats:
            program_stats[program_code] = {
                'promoted': 0, 'retained': 0, 'pending': 0, 'other': 0,
                'grades': [], 'total': 0
            }
        program_stats[program_code]['total'] += 1
        if s.final_status == 'promoted':
            program_stats[program_code]['promoted'] += 1
        elif s.final_status == 'retained':
            program_stats[program_code]['retained'] += 1
        elif s.final_status == 'pending':
            program_stats[program_code]['pending'] += 1
        else:
            program_stats[program_code]['other'] += 1
        if s.overall_grade:
            program_stats[program_code]['grades'].append(float(s.overall_grade))

        # Section grouping
        section_name = s.section.name if s.section else 'Unknown'
        section_key = f"{program_code} — {section_name}"
        if section_key not in section_stats:
            section_stats[section_key] = {
                'section': section_name,
                'program': program_code,
                'grade_level': s.grade_level.name if s.grade_level else '',
                'promoted': 0, 'retained': 0, 'pending': 0,
                'grades': [], 'total': 0
            }
        section_stats[section_key]['total'] += 1
        if s.final_status == 'promoted':
            section_stats[section_key]['promoted'] += 1
        elif s.final_status == 'retained':
            section_stats[section_key]['retained'] += 1
        elif s.final_status == 'pending':
            section_stats[section_key]['pending'] += 1
        if s.overall_grade:
            section_stats[section_key]['grades'].append(float(s.overall_grade))

    # Individual student records
    individual = []
    for s in statuses:
        sd = getattr(s.student, 'student_data', None)
        individual.append({
            'lrn': s.student.lrn,
            'name': f"{sd.last_name}, {sd.first_name}" if sd else s.student.lrn,
            'program': s.section.program.code if s.section and s.section.program else '—',
            'section': s.section.name if s.section else '—',
            'grade_level': s.grade_level.name if s.grade_level else '—',
            'overall_grade': float(s.overall_grade) if s.overall_grade else None,
            'final_status': s.final_status,
        })

    # Compute averages
    for k in program_stats:
        g = program_stats[k]['grades']
        program_stats[k]['average'] = round(sum(g) / len(g), 2) if g else None
        del program_stats[k]['grades']

    for k in section_stats:
        g = section_stats[k]['grades']
        section_stats[k]['average'] = round(sum(g) / len(g), 2) if g else None
        del section_stats[k]['grades']

    return JsonResponse({
        'success': True,
        'school_year': school_year.year_label,
        'summary': {
            'total': total,
            'promoted': promoted,
            'retained': retained,
            'incomplete': incomplete,
            'others': others,
        },
        'by_program': [
            {'program': k, **v} for k, v in sorted(program_stats.items())
        ],
        'by_section': [
            {'key': k, **v} for k, v in sorted(section_stats.items())
        ],
        'individual': sorted(individual, key=lambda x: x['name']),
    })


@admin_required
def archive_probation(request):
    """Tab 4 — Probation records for a given school year."""
    sy_id = request.GET.get('sy')
    if not sy_id:
        return JsonResponse({'success': False, 'error': 'sy parameter required.'}, status=400)

    try:
        school_year = SchoolYear.objects.get(id=sy_id, is_active=False)
    except SchoolYear.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'School year not found.'}, status=404)

    records = ProbationRecord.objects.filter(
        school_year=school_year
    ).select_related(
        'student', 'student__student_data',
        'grade_level', 'reinstated_by'
    ).order_by('-flagged_at')

    data = []
    for rec in records:
        sd = getattr(rec.student, 'student_data', None)
        data.append({
            'lrn': rec.student.lrn,
            'name': f"{sd.last_name}, {sd.first_name}" if sd else rec.student.lrn,
            'grade_level': rec.grade_level.name if rec.grade_level else '—',
            'previous_program': rec.previous_program,
            'moved_to_program': rec.moved_to_program,
            'reason': rec.reason,
            'is_active': rec.is_active,
            'flagged_at': rec.flagged_at.strftime('%b %d, %Y') if rec.flagged_at else '—',
            'lifted_by': (
                rec.reinstated_by.get_full_name() or rec.reinstated_by.username
            ) if rec.reinstated_by else None,
            'lifted_at': rec.reinstated_at.strftime('%b %d, %Y') if rec.reinstated_at else None,
            'lift_reason': rec.reinstatement_reason or None,
        })

    return JsonResponse({
        'success': True,
        'school_year': school_year.year_label,
        'total': len(data),
        'active': sum(1 for d in data if d['is_active']),
        'lifted': sum(1 for d in data if not d['is_active']),
        'records': data,
    })


# ─────────────────────────────────────────────────────────────
# EXPORTS
# ─────────────────────────────────────────────────────────────

def _header_style():
    fill = PatternFill(start_color='991B1B', end_color='991B1B', fill_type='solid')
    font = Font(bold=True, color='FFFFFF')
    align = Alignment(horizontal='center', vertical='center')
    return fill, font, align


@admin_required
def export_archive_enrollment(request):
    """Export enrollment summary to Excel."""
    sy_id = request.GET.get('sy')
    try:
        school_year = SchoolYear.objects.get(id=sy_id, is_active=False)
    except SchoolYear.DoesNotExist:
        return HttpResponse('School year not found.', status=404)

    enrollments = StudentEnrollment.objects.filter(school_year=school_year)
    program_selections = ProgramSelection.objects.filter(school_year=school_year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Enrollment Summary'

    fill, font, align = _header_style()

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f'Enrollment Summary — {school_year.year_label}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 24

    # Overall summary
    ws.append([])
    ws.append(['Overall Summary'])
    ws['A3'].font = Font(bold=True, size=11)
    ws.append(['Total Enrolled', enrollments.count()])
    ws.append(['New', enrollments.filter(enrollee_type='new').count()])
    ws.append(['Continuing', enrollments.filter(enrollee_type='continuing').count()])
    ws.append(['Transferee', enrollments.filter(enrollee_type='transferee').count()])
    ws.append(['Approved', enrollments.filter(enrollment_status='approved').count()])
    ws.append(['Rejected', enrollments.filter(enrollment_status='rejected').count()])

    ws.append([])
    ws.append(['Per Program Breakdown'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)

    headers = ['Program', 'Total', 'Approved', 'Rejected', 'Pending']
    ws.append(headers)
    header_row = ws.max_row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align

    programs = program_selections.values('selected_program_code').annotate(
        total=Count('student'),
        approved=Count('student', filter=Q(admin_approved=True)),
        rejected=Count('student', filter=Q(admin_rejected=True)),
    ).order_by('selected_program_code')

    for p in programs:
        if p['selected_program_code']:
            pending = p['total'] - p['approved'] - p['rejected']
            ws.append([p['selected_program_code'], p['total'], p['approved'], p['rejected'], pending])

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="archive_enrollment_{school_year.year_label}.xlsx"'
    wb.save(response)
    return response


@admin_required
def export_archive_sections(request):
    """Export all sections and their student lists to Excel."""
    sy_id = request.GET.get('sy')
    try:
        school_year = SchoolYear.objects.get(id=sy_id, is_active=False)
    except SchoolYear.DoesNotExist:
        return HttpResponse('School year not found.', status=404)

    sections = Section.objects.filter(school_year=school_year).select_related(
        'program', 'grade_level', 'adviser'
    ).order_by('program__code', 'grade_level__code', 'name')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    fill, font, align = _header_style()

    for section in sections:
        ws_name = f"{section.program.code if section.program else 'UNK'}-{section.name}"[:31]
        ws = wb.create_sheet(title=ws_name)

        ws.merge_cells('A1:F1')
        ws['A1'] = f"{section.name} — {school_year.year_label}"
        ws['A1'].font = Font(bold=True, size=13)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.append([
            'Program:', section.program.code if section.program else '—',
            'Grade Level:', section.grade_level.name if section.grade_level else '—',
            'Adviser:', section.adviser.get_full_name() if section.adviser else '—'
        ])
        ws.append([])

        headers = ['#', 'LRN', 'Name', 'Gender', 'Overall Grade', 'Promotion Status']
        ws.append(headers)
        h_row = ws.max_row
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=h_row, column=col)
            c.fill = fill
            c.font = font
            c.alignment = align

        students = ProgramSelection.objects.filter(
            assigned_section=section, admin_approved=True
        ).select_related('student', 'student__student_data').order_by(
            'student__student_data__last_name'
        )

        for idx, ps in enumerate(students, 1):
            sd = getattr(ps.student, 'student_data', None)
            try:
                status = StudentAcademicYearStatus.objects.get(
                    student=ps.student, school_year=school_year
                )
                overall = float(status.overall_grade) if status.overall_grade else '—'
                final_status = status.final_status.title()
            except StudentAcademicYearStatus.DoesNotExist:
                overall = '—'
                final_status = '—'

            ws.append([
                idx,
                ps.student.lrn,
                f"{sd.last_name}, {sd.first_name}" if sd else ps.student.lrn,
                sd.gender.capitalize() if sd else '—',
                overall,
                final_status,
            ])

        for col_letter, width in zip(['A','B','C','D','E','F'], [5, 15, 30, 10, 15, 18]):
            ws.column_dimensions[col_letter].width = width

    if not wb.sheetnames:
        wb.create_sheet('No Data')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="archive_sections_{school_year.year_label}.xlsx"'
    wb.save(response)
    return response


@admin_required
def export_archive_academic(request):
    """Export academic performance to Excel."""
    sy_id = request.GET.get('sy')
    try:
        school_year = SchoolYear.objects.get(id=sy_id, is_active=False)
    except SchoolYear.DoesNotExist:
        return HttpResponse('School year not found.', status=404)

    statuses = StudentAcademicYearStatus.objects.filter(
        school_year=school_year
    ).select_related(
        'student', 'student__student_data',
        'section', 'section__program', 'grade_level'
    ).order_by('section__program__code', 'section__name', 'student__student_data__last_name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Academic Performance'
    fill, font, align = _header_style()

    ws.merge_cells('A1:G1')
    ws['A1'] = f'Academic Performance — {school_year.year_label}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 24
    ws.append([])

    headers = ['#', 'LRN', 'Student Name', 'Program', 'Section', 'Overall Grade', 'Status']
    ws.append(headers)
    h_row = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=h_row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = align

    for idx, s in enumerate(statuses, 1):
        sd = getattr(s.student, 'student_data', None)
        ws.append([
            idx,
            s.student.lrn,
            f"{sd.last_name}, {sd.first_name}" if sd else s.student.lrn,
            s.section.program.code if s.section and s.section.program else '—',
            s.section.name if s.section else '—',
            float(s.overall_grade) if s.overall_grade else '—',
            s.final_status.title(),
        ])

    for col_letter, width in zip(['A','B','C','D','E','F','G'], [5, 15, 30, 10, 20, 15, 15]):
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="archive_academic_{school_year.year_label}.xlsx"'
    wb.save(response)
    return response


@admin_required
def export_archive_probation(request):
    """Export probation records to Excel."""
    sy_id = request.GET.get('sy')
    try:
        school_year = SchoolYear.objects.get(id=sy_id, is_active=False)
    except SchoolYear.DoesNotExist:
        return HttpResponse('School year not found.', status=404)

    records = ProbationRecord.objects.filter(
        school_year=school_year
    ).select_related(
        'student', 'student__student_data', 'grade_level', 'reinstated_by'
    ).order_by('-flagged_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Probation Records'
    fill, font, align = _header_style()

    ws.merge_cells('A1:H1')
    ws['A1'] = f'STE Probation Records — {school_year.year_label}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 24
    ws.append([])

    headers = ['#', 'LRN', 'Student Name', 'Grade Level', 'Reason', 'Status', 'Lifted By', 'Lifted At']
    ws.append(headers)
    h_row = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=h_row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = align

    for idx, rec in enumerate(records, 1):
        sd = getattr(rec.student, 'student_data', None)
        lifted_by = ''
        if rec.reinstated_by:
            lifted_by = rec.reinstated_by.get_full_name() or rec.reinstated_by.username

        ws.append([
            idx,
            rec.student.lrn,
            f"{sd.last_name}, {sd.first_name}" if sd else rec.student.lrn,
            rec.grade_level.name if rec.grade_level else '—',
            rec.reason,
            'Active' if rec.is_active else 'Lifted',
            lifted_by,
            rec.reinstated_at.strftime('%b %d, %Y') if rec.reinstated_at else '—',
        ])

    for col_letter, width in zip(['A','B','C','D','E','F','G','H'], [5,15,30,12,50,10,25,15]):
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="archive_probation_{school_year.year_label}.xlsx"'
    wb.save(response)
    return response