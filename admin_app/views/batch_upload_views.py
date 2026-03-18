from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
from admin_app.models import Teacher, Position, Department
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import io
import pandas as pd
from datetime import datetime


@login_required
def download_teacher_template(request):
    """Download Excel template for batch teacher upload"""
    try:
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Teacher Template"

        # Add headers
        headers = [
            'first_name',
            'middle_name',
            'last_name',
            'email',
            'position_id',
            'department_id',
            'address'
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(
            start_color='991B1B', end_color='991B1B', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = max(
                len(header) + 4, 15)

        # Add sample data row
        sample_row = [
            'John',
            'M',
            'Doe',
            'john.doe@example.com',
            '1',  # position_id (you'll need to replace with actual IDs)
            '1',  # department_id
            'Sample Address'
        ]
        ws.append(sample_row)

        # Add second sheet with available positions and departments
        ws2 = wb.create_sheet("Reference Data")
        ws2.append(["Available Positions", "Department IDs"])

        # Add positions
        positions = Position.objects.filter(
            is_active=True).values_list('id', 'name')
        for pos_id, pos_name in positions:
            ws2.append([f"{pos_id}: {pos_name}", ""])

        ws2.append([])
        ws2.append(["Available Departments"])
        departments = Department.objects.filter(
            is_active=True).values_list('id', 'name')
        for dept_id, dept_name in departments:
            ws2.append([f"{dept_id}: {dept_name}", ""])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="teacher_template_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)

        return response

    except Exception as e:
        messages.error(request, f'Error generating template: {str(e)}')
        return redirect('admin_app:settings')


@login_required
def batch_upload_teachers(request):
    """Handle batch upload of teachers via Excel/CSV"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        # Check if file was uploaded
        if 'upload_file' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'No file uploaded'})

        uploaded_file = request.FILES['upload_file']
        print(
            f"File received: {uploaded_file.name}, size: {uploaded_file.size}")

        # Parse file based on extension
        file_name = uploaded_file.name.lower()
        try:
            if file_name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            elif file_name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(uploaded_file)
            else:
                return JsonResponse({'success': False, 'message': 'Invalid file format. Please upload .xlsx, .xls, or .csv file'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error reading file: {str(e)}'})

        print(f"DataFrame loaded: {len(df)} rows")
        print(f"Columns found: {list(df.columns)}")

        # Validate required columns
        required_columns = ['first_name', 'last_name', 'email']
        missing_columns = [
            col for col in required_columns if col not in df.columns]

        if missing_columns:
            return JsonResponse({
                'success': False,
                'message': f'Missing required columns: {", ".join(missing_columns)}. Found columns: {", ".join(list(df.columns))}'
            })

        # Process records
        results = {
            'total': len(df),
            'success': 0,
            'failed': 0,
            'errors': []
        }

        for index, row in df.iterrows():
            try:
                # Extract data with defaults
                first_name = str(row.get('first_name', '')).strip(
                ) if pd.notna(row.get('first_name')) else ''
                middle_name = str(row.get('middle_name', '')).strip(
                ) if pd.notna(row.get('middle_name')) else None
                last_name = str(row.get('last_name', '')).strip(
                ) if pd.notna(row.get('last_name')) else ''
                email = str(row.get('email', '')).strip(
                ).lower() if pd.notna(row.get('email')) else ''
                address = str(row.get('address', '')).strip(
                ) if pd.notna(row.get('address')) else None

                # Handle position_id
                position_id = None
                if pd.notna(row.get('position_id')):
                    try:
                        pos_str = str(row.get('position_id')).strip()
                        if ':' in pos_str:
                            position_id = int(pos_str.split(':')[0].strip())
                        else:
                            position_id = int(float(pos_str))
                    except (ValueError, TypeError):
                        # Skip position if invalid
                        pass

                # Handle department_id
                department_id = None
                if pd.notna(row.get('department_id')):
                    try:
                        dept_str = str(row.get('department_id')).strip()
                        if ':' in dept_str:
                            department_id = int(dept_str.split(':')[0].strip())
                        else:
                            department_id = int(float(dept_str))
                    except (ValueError, TypeError):
                        # Skip department if invalid
                        pass

                # Validate required fields
                if not first_name:
                    raise ValidationError("First name is required")
                if not last_name:
                    raise ValidationError("Last name is required")
                if not email:
                    raise ValidationError("Email is required")

                # Check email uniqueness
                if Teacher.objects.filter(email__iexact=email).exists():
                    raise ValidationError(
                        f"Teacher with email '{email}' already exists")

                # Get position and department objects if IDs provided
                position = None
                if position_id:
                    try:
                        position = Position.objects.get(
                            id=position_id, is_active=True)
                    except Position.DoesNotExist:
                        results['errors'].append(
                            f"Row {index + 2}: Position ID {position_id} not found - continuing without position")

                department = None
                if department_id:
                    try:
                        department = Department.objects.get(
                            id=department_id, is_active=True)
                    except Department.DoesNotExist:
                        results['errors'].append(
                            f"Row {index + 2}: Department ID {department_id} not found - continuing without department")

                # Create teacher
                teacher = Teacher.objects.create(
                    first_name=first_name,
                    middle_name=middle_name,
                    last_name=last_name,
                    email=email,
                    address=address,
                    position=position,
                    department=department
                )

                print(
                    f"Created teacher: {teacher.id} - {teacher.get_full_name()}")
                results['success'] += 1

            except Exception as e:
                results['failed'] += 1
                error_msg = f"Row {index + 2}: {str(e)}"
                results['errors'].append(error_msg)
                print(f"Error: {error_msg}")

        # Prepare response message
        if results['failed'] == 0:
            message = f'Successfully uploaded {results["success"]} teacher(s)!'
            return JsonResponse({
                'success': True,
                'message': message,
                'data': results
            })
        elif results['success'] > 0:
            message = f'Uploaded {results["success"]} teacher(s) with {results["failed"]} error(s).'
            return JsonResponse({
                'success': True,
                'message': message,
                'data': results
            })
        else:
            message = f'Upload failed. {results["failed"]} error(s).'
            return JsonResponse({
                'success': False,
                'message': message,
                'data': results
            })

    except Exception as e:
        print(f"Batch upload error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': f'Error processing file: {str(e)}',
            'data': {'errors': [str(e)]}
        })
